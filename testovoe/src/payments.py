from aiogram.types import LabeledPrice
from main import dp
from aiogram import Bot, types
from local_settings import PAYMENT_SBERBANK
from src.app import bot
from testovoeapi.models import Order
from channels.db import database_sync_to_async
from aiogram.types import ContentTypes
from aiogram.utils.callback_data import CallbackData
from testovoeapi.models import Cart, CartItem
from aiogram.dispatcher import FSMContext

from openpyxl import Workbook
from openpyxl import load_workbook

import uuid
import openpyxl
from openpyxl.utils import get_column_letter
from states import CheckoutStates


order_checkout_cd = CallbackData("order_checkout", "user_id")
checkout_cd = CallbackData("checkout", "user_id")


@database_sync_to_async
def update_order_status(order_id, status):
    order = Order.objects.get(order_id=order_id)
    order.status = status
    order.save()


async def send_invoice(chat_id, order_id, amount, currency='RUB'):
    prices = [LabeledPrice(label='Your Product', amount=int(amount * 100))]
    payload = f"{order_id}"
    # Convert to the smallest currency unit (kopecks)
    amount = max(1, int(amount * 100))  # Convert to the smallest currency unit (kopecks)
    amount = min(amount, 15000 * 100)


    await bot.send_invoice(
        chat_id=chat_id,
        title='Your Product',
        description='Description of your product',
        provider_token=PAYMENT_SBERBANK,
        start_parameter='start_parameter',
        currency=currency,
        prices=prices,
        payload=payload
    )


@dp.pre_checkout_query_handler()
async def process_pre_checkout_query(pre_checkout_query: types.PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True)


@dp.message_handler(content_types=ContentTypes.SUCCESSFUL_PAYMENT)
async def process_successful_payment(message: types.Message):
    successful_payment = message.successful_payment
    order_id = successful_payment.invoice_payload


    await update_order_status(order_id, "completed")

    order = await get_order_by_id(order_id)
    name = order.name
    phone = order.phone
    total_amount = order.total_amount


    save_order_to_excel(order_id, name, phone, total_amount)

    await message.answer("Payment successful!")

@database_sync_to_async
def get_order_by_id(order_id):
    order = Order.objects.get(order_id=order_id)
    return order

@dp.callback_query_handler(checkout_cd.filter())
async def checkout(query: types.CallbackQuery, callback_data: dict):
    user_id = int(callback_data["user_id"])

    await CheckoutStates.entering_name.set()
    await query.message.answer("Пожалуйста, введите ваше имя:")




@database_sync_to_async
def create_order(user_id, name, phone):

    cart = Cart.objects.get(user_id=user_id)
    cart_items = CartItem.objects.filter(cart=cart)


    total_amount = sum(item.product.price * item.quantity for item in cart_items)
    order_id = str(uuid.uuid4())  # Generate a unique order_id using UUID
    order = Order(user_id=user_id, cart=cart, total_amount=total_amount, order_id=order_id, status="pending",name=name, phone=phone)

    order.save()

    return order.order_id, total_amount


@dp.message_handler(lambda message: not message.text.startswith('/'), state=CheckoutStates.entering_name)
async def process_name(message: types.Message, state: FSMContext):
    name = message.text
    await state.update_data(name=name)

    await CheckoutStates.entering_phone.set()
    await message.answer("Пожалуйста, введите ваш номер телефона:")


@dp.message_handler(lambda message: not message.text.startswith('/'), state=CheckoutStates.entering_phone)
async def process_phone(message: types.Message, state: FSMContext):
    phone = message.text
    await state.update_data(phone=phone)

    await CheckoutStates.entering_other_info.set()
    user_data = await state.get_data()
    name = user_data["name"]

    user_id = message.from_user.id


    order_id, amount = await create_order(user_id, name, phone)


    await send_invoice(chat_id=user_id, order_id=order_id, amount=amount)

    await message.answer("Заказ создан. Ожидаем оплату.")

    await state.finish()
#EXCEL

def save_order_to_excel(order_id, name, phone, total_amount):
    filename = 'orders.xlsx'

    try:
        # Если файл уже существует, открываем его, иначе создаем новый
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

        # Создаем строку заголовка для нового файла
        header_row = ["Order ID", "Name", "Phone", "Total Amount"]
        sheet.append(header_row)

    # Получаем активный лист
    sheet = workbook.active

    # Заполняем ячейки данными о заказе
    new_row = [order_id, name, phone, total_amount]
    sheet.append(new_row)

    # Сохраняем изменения в файле
    workbook.save(filename)